#!/usr/bin/env python
# coding: utf-8

import yadisk
from tqdm import tqdm 
from openpyxl import load_workbook
import os
import pandas as pd
import openpyxl
import re
from matplotlib import pyplot as plt
import plotly.express as px
from dotenv import load_dotenv

# это функция для создания столбца с названиями ФО, проходимся циклом построчно и заполняем строки 
# удаляем строки из столбца Регион содержащие "ФО" или "РОССИЯ"
# перемещаем столбец с ФО на 1 меcто по индексу в таблице
def regions(table):
    table['ФО'] = ''
    current_value = None
    for i, row in table.iterrows():
        if isinstance(row['Регион'], str) and 'ФО' in row['Регион']:
            current_value = row['Регион']
        table.at[i, 'ФО'] = current_value
    table = table[table['Регион'].str.contains('РОССИЯ') == False]
    table = table[table['Регион'].str.contains('ФО') == False]
    column_to_move = table['ФО']
    table = table.drop(columns=['ФО'])
    table.insert(1, 'ФО', column_to_move)
    return table

# Подключаемся к Яндекс диску
dotenv_path = '/Users/daria/Desktop/my_env/untitled.env'
load_dotenv(dotenv_path=dotenv_path)
app_id = os.getenv('APP_ID')
secret_id = os.getenv('SECRET_ID')
ya_token = os.getenv('YA_TOKEN')

# создаем коннектор
y = yadisk.YaDisk(app_id, secret_id, ya_token)

# проверяем токен
print(y.check_token(), 'проверяем токен')

# Создаем список файлов для загрузки
list_of_files = []
for el in tqdm(list(y.listdir('Handler'))):
    if el['path'].endswith('.xlsx') and el['path'] not in list_of_files:
        list_of_files.append(el['path'])
        
print('Создаем список файлов для загрузки')

# проверяем, файлов должно быть 147
print(len(list_of_files), 'файлов на яндекс диске')

# создаем папку загрузки
load_path = os.getcwd() + '/handler'
print('Cоздаем папку загрузки')

# если директория не существует, то создаем ее
if not os.path.exists(load_path):
    os.mkdir(load_path)

# меняем рабочую директорию на созданную
os.chdir(load_path)
print('Если директория не существует, то создаем ее и меняем рабочую директорию на созданную')

# проверяем
print(os.getcwd(), 'Проверяем директорию')

# загружаем файлы с яндекс диска
for file in tqdm(list_of_files):
#     проверяем наличие файла в папке, чтобы лишний раз не скачивать 
    if file.split('/')[-1] not in os.listdir():
        y.download(file.split(':')[1], file.split('/')[-1])
    else:
        pass

new_list = [file for file in os.listdir() if file.endswith('.xlsx')]
print(len(new_list), 'файлов в нашей папке, с учетом финальных 4х таблиц')

# отсекаем таблицы что не выделены жирным в тз
unwanted_prefixes = ['2021_Таблица_019_Сост', '2021_Таблица_020_Сост','2021_Таблица_021_Сост', 
                     '2021_Таблица_022_Сост', '2021_Таблица_023_Сост', '2021_Таблица_057_Сост', 
                     '2021_Таблица_005_Зло', '2021_Таблица_006_Зло', '2021_Таблица_007_Зло', 
                     '2021_Таблица_065_Зло']
print('Отсекаем таблицы, что не нужны нам по тз')

working_list = [name for name in new_list if not any(name.startswith(prefix) 
                                                            for prefix in unwanted_prefixes)]
print(len(working_list), 'таблиц для работы')

tables_with_1_list = []
tables_with_2_lists = []

# проходимся по списку таблиц и разделяем таблицы по количесву листов в два списка
for file in tqdm(working_list):
    workbook = openpyxl.load_workbook(file)
    sheet_names = workbook.sheetnames
    if len(sheet_names) > 1:
        tables_with_2_lists.append(file)
    else:
        tables_with_1_list.append(file)

print(len(tables_with_1_list), 'таблиц с 1 листом')
print(len(tables_with_2_lists), 'таблиц с 2 листами')

tables_58_85 = []
tables_24_51_1 = []
tables_24_51_2 = []
tables_12_103 = []

# проходимся по таблицам с 1 листом, таблицы что имеют в названии "Сост" и 10 столбцов добавляются в список 
# tables_58_85, 
# выводятся на экран таблицы, которые имеют в названии "Сост" и не 10 столбцов,
# таблицы, что имеют в названии "Злокач" добавляются в список tables_12_103
for name in tqdm(tables_with_1_list):
    if 'Сост' in name: 
        df = pd.read_excel(name)
        num_columns = len(df.columns)
        if num_columns == 10:
            tables_58_85.append(name)
        else:
            print(name, 'эта таблица не подходит по условиям для дальнейшей работы')
    elif 'Злокач' in name:
        tables_12_103.append(name)
print(len(tables_58_85), 'Таблиц в списке tables_58_85')
print(len(tables_12_103), 'Таблиц в списке tables_12_103')

# пути таблиц в списках tables_24_51_1 и tables_24_51_2 одинаковы, но разные листы, 
# пока что нам нужны только пути - добавляем их
tables_24_51_1 = tables_with_2_lists
tables_24_51_2 = tables_with_2_lists
print(len(tables_24_51_1), 'таблиц в списке tables_24_51_1')
print(len(tables_24_51_2), 'таблиц в списке tables_24_51_2')

# списки с таблицами для работы готовы
final_one = []
print('Работаем с первой группой таблиц, 58-85. Извлекаем необходимые данные')

# Проходимся циклом по первой группе таблиц, переименовываем столбцы, 
# и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц
for element in tqdm(tables_58_85):
    all_first_header = load_workbook(element).active['A1'].value.split('  ')
    ind, loc = ' '.join(all_first_header[0].split('\n')[0:2]).capitalize(), all_first_header[0].split('\n')[2]
    table = all_first_header[-1].strip()
    year = re.search(r'\b(\d{4})\b', all_first_header[0]).group(1)
    df = pd.read_excel(element, header=[2])
    df.columns = ['Регион',
                  'Число ЗНО, выявленных в отчетном году, радикальное лечение которых закончено в отчетном году, чел.',
                  'Число ЗНО, выявленных в отчетном году, радикальное лечение которых закончено в отчетном году, % от впервые выявленных',
                  'Число ЗНО, выявленных в отчетном году, радикальное лечение которых будет продолжено в отчетом году, чел.',
                  'Число ЗНО, выявленных в отчетном году, радикальное лечение которых будет продолжено в отчетном году, % от впервые выявленных',
                  'В том числе с использованием методов только хирургического, %',
                  'В том числе с использованием методов только лучевого, %',
                  'В том числе с использованием методов только лекарственного, %',
                  'В том числе с использованием методов комбинированного или комплексного (кроме химиолучевого), %',
                  'В том числе с использованием методов химиолучевого, %']
    df['Индикатор'] = ind
    df['Локализация'] = loc
    df['Таблица'] = table
    df['Год'] = year
    final_one.append(df)
print('Проходимся циклом по первой группе таблиц, переименовываем столбцы, и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц')

# соединяем таблицы в длину
final_58_85 = pd.concat(final_one)

# применяем функцию, создающую столбец ФО
final_58_85 = regions(final_58_85)


final_58_85 = final_58_85.reset_index(drop=True)
print('Первая таблица final_58_85 - готова')

final_two = []
print('Работаем со второй группой таблиц - tables_24_51_1')

# Проходимся циклом по второй группе таблиц, переименовываем столбцы, 
# и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц
for element in tqdm(tables_24_51_1):
    all_first_header = load_workbook(element).active['A1'].value.split('   ')
    ind = ' '.join(all_first_header[0].split('\n')[0:1]).capitalize().replace('в 2021 г.', '').strip() 
    table = next((match.group() for item in all_first_header if (match := re.search(r'Таблица\s\d+|Продолжение таблицы\s\d+', item))), None)
    loc_1 = ' '.join(all_first_header[-1].split('\n')[1:2]).capitalize()
    loc_2 = ' '.join(all_first_header[0].split('\n')[0:2]).capitalize().split('. ')[-1]
    year = re.search(r'\b(\d{4})\b', all_first_header[0]).group(1)
    dff = pd.read_excel(element, sheet_name=0, header=[2])
    dff.columns = ['Регион',
                  'Взято на учет больных с впервые в жизни уст. диагнозом ЗНО',
                  'в т.ч. выявлены активно, %',
                  'Находились на учете на конец года, абсолютное число',
                  'Находились на учете на конец года, на 100 тыс. населения',
                  'из них 5 лет и более, абсолютное число',
                  'из них 5 лет и более, % от сост. на учете',
                  'Индекс накопления контингентов',
                  'Летальность, %']
    dff['Индикатор'] = ind
    dff['Таблица'] = table
    if dff['Таблица'].isin(['Таблица 33', 'Таблица 36', 'Таблица 48', 'Таблица 38', 'Таблица 43', 
                            'Таблица 35', 'Таблица 34', 'Таблица 37', 'Таблица 44']).any():
        dff['Локализация'] = loc_1
    else:
        dff['Локализация'] = loc_2
    dff['Год'] = year
    final_two.append(dff)
print('Проходимся циклом по второй группе таблиц, переименовываем столбцы, и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц')

# соединяем таблицы в длину
final_24_51_1 = pd.concat(final_two)

# применяем функцию, для создания столбца ФО и удаления лишних строк
final_24_51_1 = regions(final_24_51_1)
final_24_51_1 = final_24_51_1.reset_index(drop=True)
print('Вторая таблица final_24_51_1 - готова')

final_three = []
print('Работаем с третьей группой таблиц tables_24_51_2')

# Проходимся циклом по третьей группе таблиц, переименовываем столбцы, 
# и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц
for element in tqdm(tables_24_51_2):
    all_first_header = load_workbook(element)[load_workbook(element).sheetnames[1]]['A1'].value.split('   ')
    ind = ' '.join(all_first_header[0].split('\n')[0:1]).capitalize().replace('в 2021 г.', '').strip() 
    loc_1 = ' '.join(all_first_header[-1].split('\n')[1:2]).capitalize()
    loc_2 = ' '.join(all_first_header[0].split('\n')[0:2]).capitalize().split('. ')[-1]
    table = next((match.group() for item in all_first_header if (match := re.search(r'Таблица\s\d+|Продолжение таблицы\s\d+', item))), None)
    year = re.search(r'\b(\d{4})\b', all_first_header[0]).group(1)
    dff = pd.read_excel(element, sheet_name=1, header=[3])
    dff.columns = ['Регион',
                  'Зарегистрировано ЗНО (без учтенных посмертно)',
                  'из них, диагноз подтвержден морфологически, %',
                  'из них, имели стадию заболевания.1, %',
                  'из них, имели стадию заболевания.2, %',
                  'из них, имели стадию заболевания.3, %',
                  'из них, имели стадию заболевания.4, %',
                  'из них, имели стадию заболевания, не установлена',
                  'Летальность на первом году с момента уст. диагноза, %']
    dff['Индикатор'] = ind
    dff['Таблица'] = table
    if dff['Таблица'].isin(['Таблица 33', 'Таблица 36', 'Таблица 38', 'Таблица 35', 
                            'Таблица 34', 'Таблица 37', 'Таблица 44']).any():
        dff['Локализация'] = loc_1
    else:
        dff['Локализация'] = loc_2
    dff['Год'] = year
    final_three.append(dff)

print('Проходимся циклом по третьей группе таблиц, переименовываем столбцы, и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц')

# соединяем таблицы в длину
final_24_51_2 = pd.concat(final_three)

# применяем функцию, для создания столбца ФО и удаления лишних строк
final_24_51_2 = regions(final_24_51_2)
final_24_51_2 = final_24_51_2.reset_index(drop=True)
print('Третья таблица final_24_51_2 - готова')

final_four = []
final_five = []
print('Работаем с четвертой группой таблиц - tables_12_103')

final_four_list = []
final_five_list = []

# проходимся циклом по четвертой группе таблиц, переименовываем столбцы, и добавляем новые необходимые нам столбцы 
# с информацией из первых строк таблиц
# если в датафрейме 5 столбцов - присваеиваем одни названия колонок, если 13 - другие. 
# на выходе имеем два списка с таблицами по с 5  и 13 столбцов
for element in tqdm(tables_12_103):
    ind = load_workbook(element).active['A2'].value
    loc_1 = ' '.join(load_workbook(element).active['A4'].value.split(' ')[1:6])
    loc_2 = load_workbook(element).active['B4'].value
    table = load_workbook(element).active['A1'].value
    year = load_workbook(element).active['B3'].value
    df_f = pd.read_excel(element, header=6)
    num_columns = len(df_f.columns)
    if num_columns == 5:
        df_f.columns = ['Регион', 
                        'Абсолютное Число', 
                        'Показатель на 100 тыс. населения/"грубый"',
                        'Показатель на 100 тыс. населения/стандартизованный', 
                        'Показатель на 100 тыс. населения/ошибка']
        df_f['Индикатор'] = ind
        df_f['Локализация'] = loc_2
        df_f['Таблица'] = table
        df_f['Год'] = year
        final_four_list.append(df_f)
    elif num_columns == 13:
        df_f.columns = ['Регион',
                       'Все население/Абсолютное Число',
                       'Все население/Показатель на 100 тыс. населения/"грубый"',
                       'Все население/Показатель на 100 тыс. населения/стандартизованный',
                       'Все население/Показатель на 100 тыс. населения/ошибка',
                       'Мужчины/Абсолютное Число',
                       'Мужчины/Показатель на 100 тыс. населения/"грубый"',
                       'Мужчины/Показатель на 100 тыс. населения/стандартизованный',
                       'Мужчины/Показатель на 100 тыс. населения/ошибка',
                       'Женщины/Абсолютное Число',
                       'Женщины/Показатель на 100 тыс. населения/"грубый"',
                       'Женщины/Показатель на 100 тыс. населения/стандартизованный',
                       'Женщины/Показатель на 100 тыс. населения/ошибка']
        df_f['Индикатор'] = ind
        df_f['Таблица'] = table
        if df_f['Таблица'].isin(['Таблица 12','Таблица 51', 'Таблица 20', 'Таблица 71']).any():
            df_f['Локализация'] = loc_1
        else:
            df_f['Локализация'] = loc_2
        if df_f['Таблица'].isin(['Таблица 12']).any():
            df_f['Год'] = load_workbook(element).active['A3'].value.split(' ')[1]    
        else:
            df_f['Год'] = year
        final_five_list.append(df_f)
    else:
        print(element)
        
print('Проходимся циклом по четвертой группе таблиц, переименовываем столбцы, и добавляем новые необходимые нам столбцы с информацией из первых строк таблиц, если в датафрейме 5 столбцов - присваеиваем одни названия колонок, если 13 - другие. На выходе имеем два списка с таблицами по с 5  и 13 столбцов')

# Соединяем таблицы в длину 
final_four = pd.concat(final_four_list, ignore_index=True)
final_five = pd.concat(final_five_list, ignore_index=True)

# Создаем новый столбец 'Гендер' в final_four
# Применяется функция к каждой строке столбца 'Локализация'
# Если значение el содержится в первом списке (мужские органы), то присваивается 'Мужчины'. 
# Если не содержится, но содержится во втором списке (женские органы), то присваивается 'Женщины'.
final_four['Гендер'] = final_four['Локализация'].apply(lambda el: 'Мужчины' if el in [
    'Яичко (С62)', 'Половой член (С60)', 'Предстательная железа (С61)', 
    'Другие мужские половые органы (С60,62,63)'] else 'Женщины' if el in [
    'Яичник (С56)', 'Тело матки (С54)', 'Вульва (С51)', 
    'Другие женские половые органы (С51,52,57,58)', 'Плацента (С58)',
    'Шейка матки (С53)', 'Другие новообразования матки (С54,55)', 'Влагалище (С52)'] else '')
print('В датафрейме с 5 столбцами - final_four создаем новый столбец "Гендер"')

# теперь пересобираем датафрейм final_five - выбираем нужные столбцы для объединения в длину, 
# создаем новый столбец "гендер"/ переименовываем столбцы, чтобы все одинаково назывались и соединяем в длину
all_population = pd.concat([
    final_five[['Регион',
          'Все население/Абсолютное Число',
          'Все население/Показатель на 100 тыс. населения/"грубый"',
          'Все население/Показатель на 100 тыс. населения/стандартизованный',
          'Все население/Показатель на 100 тыс. населения/ошибка',
          'Индикатор',
          'Локализация',
          'Таблица',
          'Год']].assign(Гендер='Все население')
    .rename(columns=lambda x: x.replace('Все население/', '')),
    final_five[['Регион', 
          'Мужчины/Абсолютное Число',
          'Мужчины/Показатель на 100 тыс. населения/"грубый"',
          'Мужчины/Показатель на 100 тыс. населения/стандартизованный',
          'Мужчины/Показатель на 100 тыс. населения/ошибка',
          'Индикатор',
          'Локализация',
          'Таблица',
          'Год']].assign(Гендер='Мужчины')
    .rename(columns=lambda x: x.replace('Мужчины/', '')),
    final_five[['Регион', 
          'Женщины/Абсолютное Число',
          'Женщины/Показатель на 100 тыс. населения/"грубый"',
          'Женщины/Показатель на 100 тыс. населения/стандартизованный',
          'Женщины/Показатель на 100 тыс. населения/ошибка',
          'Индикатор',
          'Локализация',
          'Таблица',
          'Год']].assign(Гендер='Женщины')
    .rename(columns=lambda x: x.replace('Женщины/', ''))], ignore_index=True)
print('Пересобираем датафрейм final_five из 13 столбцов в датафрейм с 5 столбцами и добавляем столбец "Гендер"')

# соединяем получившийся датафрейм с final_four в длину
final_12_103 = pd.concat([all_population, final_four], ignore_index=True)

# применяем функцию, для создания столбца ФО и удаления лишних строк
final_12_103 = regions(final_12_103)
final_12_103 = final_12_103.reset_index(drop=True)
print('Четвертая таблица final_12_103 - готова')

# скачиваем готовые таблицы в рабочую директорию 
final_58_85.to_excel('final_58_85.xlsx', index=False)
final_24_51_1.to_excel('final_24_51_1.xlsx', index=False)
final_24_51_2.to_excel('final_24_51_2.xlsx', index=False)
final_12_103.to_excel('final_12_103.xlsx', index=False)
print('Скачиваем готовые таблицы в рабочую директорию')

